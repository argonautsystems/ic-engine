#!/usr/bin/env python3
# Copyright 2026 InvestorClaw Contributors
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.


"""
Export portfolio analysis to CSV and XLSX formats.
Includes detailed holdings, allocation, and performance summary.
Polars-native implementation — no pandas dependency.
"""

import json
import logging
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

import polars as pl

# Fix import path for schema module (in parent directory)
sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    import openpyxl  # noqa: F401
    from openpyxl.styles import Alignment, Font, PatternFill  # noqa: F401

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)


class ReportExporter:
    def __init__(self):
        self.holdings_data = {}
        self.performance_data = {}

    @staticmethod
    def _holding_value(holding: dict) -> float:
        value = holding.get("value")
        if value is not None:
            return value
        market_value = holding.get("market_value")
        if market_value is not None:
            return market_value
        return holding.get("marketValue", 0)

    @staticmethod
    def _holding_quantity(holding: dict) -> float:
        quantity = holding.get("quantity")
        if quantity is not None:
            return quantity
        shares = holding.get("shares")
        if shares is not None:
            return shares
        return holding.get("par_value", 0)

    @staticmethod
    def _holding_amount(holding: dict) -> float:
        amount = holding.get("amount")
        if amount is not None:
            return amount
        shares = holding.get("shares")
        if shares is not None:
            return shares
        return holding.get("market_value", 0)

    @staticmethod
    def _holding_unrealized_pct(holding: dict) -> float:
        unrealized_pct = holding.get("unrealized_pct")
        if unrealized_pct is not None:
            return unrealized_pct
        unrealized_gain_loss_pct = holding.get("unrealized_gain_loss_pct")
        if unrealized_gain_loss_pct is not None:
            return unrealized_gain_loss_pct
        unrealized_gain_loss_pct = holding.get("unrealizedGainLossPct")
        if unrealized_gain_loss_pct is not None:
            return unrealized_gain_loss_pct
        return holding.get("unrealizedPct", 0)

    @staticmethod
    def _summary_value(summary: dict, *keys: str, default: float = 0.0) -> float:
        for key in keys:
            value = summary.get(key)
            if value is not None:
                return value
        return default

    @classmethod
    def _summary_to_snake(cls, summary: dict) -> dict:
        total_value = cls._summary_value(
            summary,
            "total_portfolio_value",
            "total_value",
            "totalPortfolioValue",
        )
        net_worth = cls._summary_value(
            summary,
            "net_worth",
            "net_value",
            "netValue",
            default=total_value,
        )
        return {
            "equity_value": cls._summary_value(summary, "equity_value", "equityValue"),
            "bond_value": cls._summary_value(summary, "bond_value", "bondValue"),
            "cash_value": cls._summary_value(summary, "cash_value", "cashValue"),
            "margin_value": cls._summary_value(summary, "margin_value", "marginValue"),
            "total_portfolio_value": total_value,
            "net_worth": net_worth,
            "total_unrealized_gain_loss": cls._summary_value(
                summary,
                "total_unrealized_gain_loss",
                "unrealized_gl",
                "totalUnrealizedGainLoss",
            ),
        }

    def _get_summary(self) -> dict:
        """Return summary dict in snake_case, handling legacy and CDM formats."""
        cdm_summary = self.holdings_data.get("portfolio", {}).get("summary", {})
        if cdm_summary:
            return self._summary_to_snake(cdm_summary)

        compact = self.holdings_data.get("summary", {})
        if compact:
            return self._summary_to_snake(compact)

        return self._summary_to_snake({})

    def load_data(self, holdings_file: str, performance_file: Optional[str] = None) -> None:
        """Load holdings and performance data."""
        try:
            holdings_path = Path(holdings_file).expanduser().resolve()
            with open(holdings_path, "r") as f:
                self.holdings_data = json.load(f)
            logger.info(f"Loaded holdings from {holdings_path}")

            if performance_file:
                perf_path = Path(performance_file).expanduser().resolve()
                if perf_path.exists():
                    with open(perf_path, "r") as f:
                        self.performance_data = json.load(f)
                    logger.info(f"Loaded performance from {perf_path}")
                else:
                    logger.info(f"Performance file not found, skipping: {perf_path}")

        except Exception as e:
            logger.error(f"Error loading data files: {e}")
            raise

        # Normalize schema
        from ic_engine.config.schema import normalize_portfolio

        if "data" in self.holdings_data and isinstance(self.holdings_data["data"], dict):
            self.holdings_data = self.holdings_data["data"]
        if "data" in self.performance_data and isinstance(self.performance_data["data"], dict):
            self.performance_data = self.performance_data["data"]

        from ic_engine.config.schema import validate_portfolio

        self.holdings_data = normalize_portfolio(self.holdings_data)
        validate_portfolio(self.holdings_data)

    def create_equity_report(self) -> pl.DataFrame:
        """Create equity holdings report."""
        equity_holdings = self.holdings_data["portfolio"].get("equity", {})
        equity_perf = self.performance_data.get("equities", {})

        report_rows = []

        for sym, holding in equity_holdings.items():
            perf = equity_perf.get(sym, {})

            report_rows.append(
                {
                    "Asset Type": "Equity",
                    "Symbol": sym,
                    "Shares": holding.get("shares", 0),
                    "Purchase Price": holding.get("purchase_price", 0),
                    "Purchase Date": holding.get("purchase_date", "N/A"),
                    "Current Price": holding.get("current_price", 0),
                    "Current Value": self._holding_value(holding),
                    "Unrealized G/L": holding.get("unrealized_gain_loss", 0),
                    "Unrealized %": self._holding_unrealized_pct(holding),
                    "Sector": holding.get("sector", "Unknown"),
                    "Total Return %": perf.get("total_return_pct", 0),
                    "YTD Return %": perf.get("ytd_return_pct", 0),
                    "12M Return %": perf.get("rolling_12m_return_pct", 0),
                    "Dividend Income": perf.get("dividend_income", 0),
                    "Volatility 30D": perf.get("volatility_30d", 0),
                    "Beta": perf.get("beta", 0),
                    "Sharpe Ratio": perf.get("sharpe_ratio", 0),
                    "Max Drawdown %": perf.get("max_drawdown_pct", 0),
                }
            )

        return pl.DataFrame(report_rows) if report_rows else pl.DataFrame()

    def create_bond_report(self) -> pl.DataFrame:
        """Create bond holdings report."""
        bond_holdings = self.holdings_data["portfolio"].get("bond", {})

        report_rows = []

        for sym, holding in bond_holdings.items():
            report_rows.append(
                {
                    "Asset Type": "Bond",
                    "Symbol": sym,
                    "Quantity": self._holding_quantity(holding),
                    "Purchase Price": holding.get("purchase_price", 0),
                    "Purchase Date": holding.get("purchase_date", "N/A"),
                    "Current Price": holding.get("current_price", 0),
                    "Current Value": self._holding_value(holding),
                    "Unrealized G/L": holding.get("unrealized_gain_loss", 0),
                    "Unrealized %": self._holding_unrealized_pct(holding),
                    "Coupon Rate %": holding.get("coupon_rate", 0),
                    "Maturity Date": holding.get("maturity_date", "N/A"),
                    "YTM %": holding.get("ytm", "N/A"),
                    "Duration": holding.get("duration", "N/A"),
                }
            )

        return pl.DataFrame(report_rows) if report_rows else pl.DataFrame()

    def create_cash_report(self) -> pl.DataFrame:
        """Create cash holdings report."""
        cash_holdings = self.holdings_data["portfolio"].get("cash", {})

        report_rows = []

        for sym, holding in cash_holdings.items():
            report_rows.append(
                {
                    "Asset Type": "Cash",
                    "Account Name": sym,
                    "Amount": self._holding_amount(holding),
                    "Interest Rate %": holding.get("interest_rate", 0) * 100,
                    "Current Value": self._holding_value(holding),
                }
            )

        return pl.DataFrame(report_rows) if report_rows else pl.DataFrame()

    def create_margin_report(self) -> pl.DataFrame:
        """Create margin debt report."""
        margin_holdings = self.holdings_data["portfolio"].get("margin", {})

        report_rows = []

        for sym, holding in margin_holdings.items():
            report_rows.append(
                {
                    "Asset Type": "Margin",
                    "Loan ID": sym,
                    "Principal": holding.get("principal", 0),
                    "Interest Rate %": holding.get("interest_rate", 0) * 100,
                    "Interest Accrued": holding.get("interest_accrued", 0),
                    "Total Debt": holding.get("total_debt", 0),
                }
            )

        return pl.DataFrame(report_rows) if report_rows else pl.DataFrame()

    def create_summary_report(self) -> pl.DataFrame:
        """Create portfolio summary."""
        summary = self._get_summary()
        portfolio_metrics = self.performance_data.get("portfolio_metrics", {})

        rows = [
            {"Metric": "Equity Value", "Value": f"${summary['equity_value']:,.2f}"},
            {"Metric": "Bond Value", "Value": f"${summary['bond_value']:,.2f}"},
            {"Metric": "Cash Value", "Value": f"${summary['cash_value']:,.2f}"},
            {"Metric": "Margin Debt", "Value": f"${summary['margin_value']:,.2f}"},
            {"Metric": "Total Assets", "Value": f"${summary['total_portfolio_value']:,.2f}"},
            {"Metric": "Net Worth", "Value": f"${summary['net_worth']:,.2f}"},
            {
                "Metric": "Total Unrealized G/L",
                "Value": f"${summary['total_unrealized_gain_loss']:,.2f}",
            },
            {
                "Metric": "Overall Return %",
                "Value": f"{portfolio_metrics.get('overall_return_pct', 0):.2f}%",
            },
            {
                "Metric": "YTD Return %",
                "Value": f"{portfolio_metrics.get('ytd_return_pct', 0):.2f}%",
            },
            {
                "Metric": "12M Return %",
                "Value": f"{portfolio_metrics.get('rolling_12m_return_pct', 0):.2f}%",
            },
            {
                "Metric": "Average Volatility %",
                "Value": f"{portfolio_metrics.get('average_volatility', 0):.2f}%",
            },
            {"Metric": "Average Beta", "Value": f"{portfolio_metrics.get('average_beta', 0):.2f}"},
            {
                "Metric": "Average Sharpe Ratio",
                "Value": f"{portfolio_metrics.get('average_sharpe', 0):.2f}",
            },
            {
                "Metric": "Herfindahl Index",
                "Value": f"{portfolio_metrics.get('herfindahl_index', 0):.0f}",
            },
            {
                "Metric": "Diversification",
                "Value": portfolio_metrics.get("diversification", "Unknown"),
            },
            {
                "Metric": "Portfolio Last Updated",
                "Value": self.holdings_data.get("timestamp", "N/A"),
            },
        ]

        return pl.DataFrame(rows)

    def create_allocation_report(self) -> pl.DataFrame:
        """Create asset allocation report."""
        summary = self._get_summary()
        total = summary["total_portfolio_value"]

        rows = [
            {"Asset Class": "Equity", "Value": summary["equity_value"]},
            {"Asset Class": "Bonds", "Value": summary["bond_value"]},
            {"Asset Class": "Cash", "Value": summary["cash_value"]},
            {"Asset Class": "Margin Debt", "Value": summary["margin_value"]},
        ]

        df = pl.DataFrame(rows)
        if total != 0:
            df = df.with_columns(
                (pl.col("Value").abs() / total * 100).round(2).alias("Allocation %")
            )
        else:
            df = df.with_columns(pl.lit(0.0).alias("Allocation %"))

        return df

    def export_to_csv(self, output_prefix: str = "portfolio_report") -> None:
        """Export all reports to CSV files."""
        try:
            # Ensure output directory exists
            output_path = Path(output_prefix).expanduser().resolve()
            output_path.parent.mkdir(parents=True, exist_ok=True)

            equity_df = self.create_equity_report()
            bond_df = self.create_bond_report()
            cash_df = self.create_cash_report()
            margin_df = self.create_margin_report()
            summary_df = self.create_summary_report()
            allocation_df = self.create_allocation_report()

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            if not equity_df.is_empty():
                equity_file = output_path.parent / f"{output_path.stem}_equities_{timestamp}.csv"
                equity_df.write_csv(str(equity_file))
                logger.info(f"Exported equities to {equity_file}")

            if not bond_df.is_empty():
                bond_file = output_path.parent / f"{output_path.stem}_bonds_{timestamp}.csv"
                bond_df.write_csv(str(bond_file))
                logger.info(f"Exported bonds to {bond_file}")

            if not cash_df.is_empty():
                cash_file = output_path.parent / f"{output_path.stem}_cash_{timestamp}.csv"
                cash_df.write_csv(str(cash_file))
                logger.info(f"Exported cash to {cash_file}")

            if not margin_df.is_empty():
                margin_file = output_path.parent / f"{output_path.stem}_margin_{timestamp}.csv"
                margin_df.write_csv(str(margin_file))
                logger.info(f"Exported margin to {margin_file}")

            summary_file = output_path.parent / f"{output_path.stem}_summary_{timestamp}.csv"
            summary_df.write_csv(str(summary_file))
            logger.info(f"Exported summary to {summary_file}")

            allocation_file = output_path.parent / f"{output_path.stem}_allocation_{timestamp}.csv"
            allocation_df.write_csv(str(allocation_file))
            logger.info(f"Exported allocation to {allocation_file}")

        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            raise

    def export_to_excel(self, output_file: str = "portfolio_report.xlsx") -> None:
        """Export all reports to single Excel file with multiple sheets."""
        if not EXCEL_AVAILABLE:
            logger.warning(
                "openpyxl not installed. Skipping Excel export. Install with: pip install openpyxl"
            )
            return

        try:
            from openpyxl import Workbook

            # Ensure output directory exists
            output_path = Path(output_file).expanduser().resolve()
            output_path.parent.mkdir(parents=True, exist_ok=True)

            sheets = {
                "Equities": self.create_equity_report(),
                "Bonds": self.create_bond_report(),
                "Cash": self.create_cash_report(),
                "Margin": self.create_margin_report(),
                "Summary": self.create_summary_report(),
                "Allocation": self.create_allocation_report(),
            }

            wb = Workbook()
            # Remove the default empty sheet that openpyxl creates
            wb.remove(wb.active)

            for sheet_name, df in sheets.items():
                if df.is_empty():
                    continue

                ws = wb.create_sheet(title=sheet_name)

                # Write header row
                ws.append(df.columns)

                # Write data rows
                for row in df.rows():
                    ws.append(list(row))

            wb.save(str(output_path))
            logger.info(f"Exported complete portfolio report to {output_path}")

        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise


if __name__ == "__main__":
    holdings_file = sys.argv[1] if len(sys.argv) > 1 else "holdings.json"
    performance_file = sys.argv[2] if len(sys.argv) > 2 else "performance.json"
    output_format = sys.argv[3].lower() if len(sys.argv) > 3 else "both"
    output_prefix = sys.argv[4] if len(sys.argv) > 4 else "portfolio_report"

    exporter = ReportExporter()
    exporter.load_data(holdings_file, performance_file)

    if output_format in ["csv", "both"]:
        exporter.export_to_csv(output_prefix=output_prefix)

    if output_format in ["xlsx", "excel", "xls", "both"]:
        exporter.export_to_excel(output_file=f"{output_prefix}.xlsx")

    logger.info("Export complete!")
